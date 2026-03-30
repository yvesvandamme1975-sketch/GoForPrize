import openpyxl
from typing import List, Dict, Optional, Callable
from src.column_mapper import ColumnMapper


class ExcelReader:
    def __init__(self, path: str,
                 mapping_override: Optional[Dict] = None,
                 on_mapping_needed: Optional[Callable] = None):
        """
        path: .xlsx file
        mapping_override: pre-resolved mapping (from MappingDialog)
        on_mapping_needed: callback(headers, mapping) -> dict | None
                           called when required fields can't be auto-mapped
        """
        self._path = path
        self._mapping_override = mapping_override
        self._on_mapping_needed = on_mapping_needed
        self._rows: List[Dict] = []
        self._headers: List[str] = []
        self._resolved_mapping: Dict = {}
        self._load()

    def _load(self):
        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        ws = wb.active
        self._headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

        mapping = (self._mapping_override
                   or ColumnMapper.auto_map(self._headers))

        missing = ColumnMapper.missing_required(mapping)
        if missing and self._on_mapping_needed:
            result = self._on_mapping_needed(self._headers, mapping)
            if result:
                mapping = result

        self._resolved_mapping = mapping
        inv = {v: k for k, v in mapping.items() if v is not None}

        self._rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            raw = {self._headers[i]: row[i]
                   for i in range(min(len(self._headers), len(row)))}
            record: Dict = {inv.get(h, h): v for h, v in raw.items()}

            for pk in ("pvente", "ppro", "ppro_htva", "pa_htva"):
                if isinstance(record.get(pk), (int, float)):
                    record[pk] = float(record[pk])
                else:
                    try:
                        record[pk] = float(
                            str(record.get(pk, 0))
                            .replace(",", ".").replace("€", "").strip())
                    except (ValueError, TypeError):
                        record[pk] = 0.0

            for sk in ("article", "origine", "p_l"):
                record[sk] = str(record.get(sk) or "").strip()

            self._rows.append(record)
        wb.close()

    def reload(self, mapping_override=None):
        if mapping_override:
            self._mapping_override = mapping_override
        self._load()

    @property
    def headers(self) -> List[str]:
        return self._headers

    @property
    def resolved_mapping(self) -> Dict:
        return self._resolved_mapping

    def all_rows(self) -> List[Dict]:
        return list(self._rows)

    def search(self, query: str) -> List[Dict]:
        if not query:
            return self.all_rows()
        q = query.lower().strip()
        return [r for r in self._rows if q in r.get("article", "").lower()]

    def suggestions(self, query: str, limit: int = 8) -> List[str]:
        if not query:
            return []
        q = query.lower().strip()
        return [r["article"] for r in self._rows
                if q in r.get("article", "").lower()][:limit]

    def search_with_suggestions(self, query: str, limit: int = 8):
        """Single-pass: returns (suggestion_names, matching_rows)."""
        if not query:
            return [], self.all_rows()
        q = query.lower().strip()
        results = [r for r in self._rows if q in r.get("article", "").lower()]
        suggestions = [r["article"] for r in results[:limit]]
        return suggestions, results

    @staticmethod
    def format_price(value: float) -> str:
        return f"{value:.2f}".replace(".", ",")

    @staticmethod
    def format_price_per_litre(val) -> str:
        """Format price/litre: '4.77' → '4,77€/L', already formatted strings pass through."""
        s = str(val or "").strip()
        if not s:
            return ""
        if "€" in s or "/" in s or "L" in s:
            return s  # already formatted
        try:
            num = float(s.replace(",", "."))
            return f"{num:.2f}".replace(".", ",") + "€/L"
        except (ValueError, TypeError):
            return s
